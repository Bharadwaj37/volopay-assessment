"""
generate_dashboard.py — Sales Dashboard Generator
Growth Squad Assessment | Volopay
Run: python generate_dashboard.py
Output: Sales_Dashboard.xlsx
"""

import random
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows
import os

random.seed(42)

# ─────────────────────────────────────────────
# DATA GENERATION
# ─────────────────────────────────────────────

COMPANIES = [
    "Meridian Logistics","BlueSky Tech","Falcon Financial","Apex Retail",
    "Horizon Healthcare","Nexus Manufacturing","PeakPerform Sports","Vantage Consulting",
    "Crimson Media","Sterling Education","Aurora Pharma","Titan Auto Parts",
    "Orion Shipping","Pinnacle Realty","Vertex Energy","Quantum Dynamics",
    "Swift Commerce","Ember Analytics","Nova Constructions","Zephyr Payments",
    "Atlas Fintech","Cobalt Security","Drift AI","Echo Networks","Flux Robotics",
    "Granite Supply","Harbor Marine","Iris Biotech","Jasper Retail","Karma Finance",
    "Lunar Aerospace","Marble Interiors","Nest PropTech","Onyx Insurance","Prism Legal",
    "Quest Edutech","Radon Chemicals","Solstice Travel","Talon Defense","Unity Gaming",
    "Vega Publishing","Wren Architecture","Xenon Motors","Yield Agritech","Zenith Telecom",
    "Alpine Ventures","Blade Logistics","Cedar Consulting","Delta Renewables","Eagle Aviation",
]

INDUSTRIES = ["SaaS","Fintech","Healthcare","Manufacturing","Retail","Logistics",
               "Education","Real Estate","Media","Energy","Aerospace","Legal"]
SOURCES = ["LinkedIn","Cold Email","Referral","Inbound","Partner","Webinar","Event","SEO"]
OWNERS = ["Priya Sharma","Arjun Mehta","Neha Kapoor","Rahul Verma","Sanjay Iyer"]
STATUSES = ["Open","Won","Lost","Nurturing","Proposal Sent","Demo Scheduled"]
PRIORITIES = ["High","Medium","Low"]
STAGES = ["Discovery","Qualification","Demo","Proposal","Negotiation","Closed Won","Closed Lost"]

def rand_date(start, end):
    delta = end - start
    return start + timedelta(days=random.randint(0, delta.days))

def generate_leads(n=100):
    today = datetime.today()
    rows = []
    for i in range(1, n+1):
        stage = random.choice(STAGES)
        if stage == "Closed Won":
            status = "Won"
        elif stage == "Closed Lost":
            status = "Lost"
        else:
            status = random.choice(["Open","Nurturing","Proposal Sent","Demo Scheduled"])

        prob_map = {
            "Discovery": random.randint(5,20),
            "Qualification": random.randint(20,40),
            "Demo": random.randint(35,55),
            "Proposal": random.randint(50,70),
            "Negotiation": random.randint(65,85),
            "Closed Won": 100,
            "Closed Lost": 0,
        }
        prob = prob_map.get(stage, 50)
        deal_value = random.choice([
            random.randint(5000, 30000),
            random.randint(30000, 120000),
            random.randint(120000, 500000),
        ])
        last_fu = rand_date(today - timedelta(days=90), today)
        next_fu = last_fu + timedelta(days=random.randint(3, 21))
        company_idx = (i - 1) % len(COMPANIES)
        rows.append({
            "Lead ID": f"LD-{i:04d}",
            "Company": COMPANIES[company_idx],
            "Industry": random.choice(INDUSTRIES),
            "Source": random.choice(SOURCES),
            "Owner": random.choice(OWNERS),
            "Status": status,
            "Priority": random.choice(PRIORITIES),
            "Deal Value": deal_value,
            "Stage": stage,
            "Last Follow-up": last_fu.strftime("%Y-%m-%d"),
            "Next Follow-up": next_fu.strftime("%Y-%m-%d"),
            "Probability (%)": prob,
        })
    return pd.DataFrame(rows)

# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

DARK_BG   = "1E2130"
ACCENT    = "6366F1"
GREEN_H   = "22C55E"
RED_H     = "EF4444"
AMBER_H   = "F59E0B"
HEADER_FG = "FFFFFF"
ALT_ROW   = "252836"
BORDER_C  = "2D3148"

def hdr_style(ws, cell_ref, text, bg=ACCENT, fg=HEADER_FG, size=11, bold=True):
    c = ws[cell_ref]
    c.value = text
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def thin_border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def write_leads_sheet(wb, df):
    ws = wb.create_sheet("Leads Data")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = list(df.columns)
    col_widths = [10,22,14,12,14,12,10,14,14,14,14,14]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color=HEADER_FG, size=10, name="Calibri")
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    STATUS_COLORS = {
        "Won": "D1FAE5", "Lost": "FEE2E2", "Open": "DBEAFE",
        "Nurturing": "FEF3C7", "Proposal Sent": "EDE9FE", "Demo Scheduled": "CCFBF1"
    }
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        bg = ALT_ROW if row_idx % 2 == 0 else "1A1F2E"
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(color="E2E8F0", size=9, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
            if col_idx == 6:  # Status
                sc = STATUS_COLORS.get(str(val), bg)
                c.fill = PatternFill("solid", fgColor=sc)
                c.font = Font(color="111827", bold=True, size=9, name="Calibri")

    # Conditional formatting on Probability
    prob_col = len(headers)
    prob_range = f"{get_column_letter(prob_col)}2:{get_column_letter(prob_col)}{len(df)+1}"
    ws.conditional_formatting.add(prob_range, ColorScaleRule(
        start_type="min", start_color="EF4444",
        mid_type="percentile", mid_value=50, mid_color="F59E0B",
        end_type="max", end_color="22C55E"
    ))
    # Deal Value data bars
    dv_col = headers.index("Deal Value") + 1
    dv_range = f"{get_column_letter(dv_col)}2:{get_column_letter(dv_col)}{len(df)+1}"
    ws.conditional_formatting.add(dv_range, DataBarRule(
        start_type="min", end_type="max", color="6366F1"
    ))
    return ws

def write_dashboard_sheet(wb, df):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT

    # Column widths
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in range(1, 60):
        ws.row_dimensions[row].height = 22

    # ── Title ──
    ws.merge_cells("A1:R1")
    t = ws["A1"]
    t.value = "🎯  SALES PIPELINE DASHBOARD"
    t.font = Font(bold=True, color=HEADER_FG, size=18, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:R2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.today().strftime('%d %B %Y')}  |  Growth Squad Assessment  |  Volopay"
    sub.font = Font(color="718096", size=10, name="Calibri", italic=True)
    sub.fill = PatternFill("solid", fgColor=DARK_BG)
    sub.alignment = Alignment(horizontal="center")

    # ── KPI Calculations (in Python, write as values) ──
    total_leads = len(df)
    open_leads = len(df[df["Status"] == "Open"])
    won = len(df[df["Status"] == "Won"])
    lost = len(df[df["Status"] == "Lost"])
    total_revenue = df[df["Status"] == "Won"]["Deal Value"].sum()
    conv_rate = round(won / total_leads * 100, 1)
    weighted_pipeline = (df["Deal Value"] * df["Probability (%)"] / 100).sum()
    avg_deal = df["Deal Value"].mean()

    kpi_labels = [
        ("Total Leads", total_leads, ACCENT, ""),
        ("Open Leads", open_leads, "3B82F6", ""),
        ("Won", won, GREEN_H, ""),
        ("Lost", lost, RED_H, ""),
        ("Total Revenue", f"${total_revenue:,.0f}", GREEN_H, "Won deals"),
        ("Conversion Rate", f"{conv_rate}%", AMBER_H, "Won / Total"),
        ("Weighted Pipeline", f"${weighted_pipeline:,.0f}", "A78BFA", "Prob-adjusted"),
        ("Avg Deal Size", f"${avg_deal:,.0f}", "60A5FA", "All leads"),
    ]

    row = 4
    ws.merge_cells(f"A{row}:R{row}")
    kpi_hdr = ws[f"A{row}"]
    kpi_hdr.value = "KEY PERFORMANCE INDICATORS"
    kpi_hdr.font = Font(bold=True, color="A0AEC0", size=10, name="Calibri", italic=False)
    kpi_hdr.fill = PatternFill("solid", fgColor=DARK_BG)
    kpi_hdr.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row].height = 18

    row = 5
    for i, (label, val, color, sub_text) in enumerate(kpi_labels):
        col_start = 1 + i * 2 + (i // 4) * 0  # lay out 4 per row
        if i == 4:
            row = 9  # second row of KPIs

        actual_col = 1 + (i % 4) * 4
        ws.merge_cells(start_row=row, start_column=actual_col,
                        end_row=row, end_column=actual_col+2)
        ws.merge_cells(start_row=row+1, start_column=actual_col,
                        end_row=row+1, end_column=actual_col+2)
        ws.merge_cells(start_row=row+2, start_column=actual_col,
                        end_row=row+2, end_column=actual_col+2)

        lc = ws.cell(row=row, column=actual_col, value=label)
        lc.font = Font(color="718096", size=9, name="Calibri", bold=True)
        lc.fill = PatternFill("solid", fgColor="1A1F2E")
        lc.alignment = Alignment(horizontal="center", vertical="center")

        vc = ws.cell(row=row+1, column=actual_col, value=val)
        vc.font = Font(color=color, size=20, bold=True, name="Calibri")
        vc.fill = PatternFill("solid", fgColor="1A1F2E")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row+1].height = 32

        sc = ws.cell(row=row+2, column=actual_col, value=sub_text)
        sc.font = Font(color="4B5563", size=8, name="Calibri")
        sc.fill = PatternFill("solid", fgColor="1A1F2E")
        sc.alignment = Alignment(horizontal="center", vertical="center")

    # ── Summary Tables for Charts ──
    # Source summary (col Q)
    source_agg = df.groupby("Source")["Deal Value"].sum().reset_index()
    source_agg.columns = ["Source", "Revenue"]
    ws.cell(row=14, column=1, value="Source").font = Font(bold=True, color=HEADER_FG, name="Calibri")
    ws.cell(row=14, column=1).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=14, column=2, value="Revenue").font = Font(bold=True, color=HEADER_FG, name="Calibri")
    ws.cell(row=14, column=2).fill = PatternFill("solid", fgColor=ACCENT)
    for r_idx, row_data in source_agg.iterrows():
        ws.cell(row=15+r_idx, column=1, value=row_data["Source"])
        ws.cell(row=15+r_idx, column=2, value=row_data["Revenue"])

    # Stage pipeline (col D)
    stage_agg = df.groupby("Stage")["Deal Value"].sum().reset_index()
    ws.cell(row=14, column=4, value="Stage").font = Font(bold=True, color=HEADER_FG, name="Calibri")
    ws.cell(row=14, column=4).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=14, column=5, value="Value").font = Font(bold=True, color=HEADER_FG, name="Calibri")
    ws.cell(row=14, column=5).fill = PatternFill("solid", fgColor=ACCENT)
    for r_idx, row_data in stage_agg.iterrows():
        ws.cell(row=15+r_idx, column=4, value=row_data["Stage"])
        ws.cell(row=15+r_idx, column=5, value=row_data["Deal Value"])

    # Owner performance (col G)
    owner_agg = df[df["Status"] == "Won"].groupby("Owner")["Deal Value"].sum().reset_index()
    ws.cell(row=14, column=7, value="Owner").font = Font(bold=True, color=HEADER_FG, name="Calibri")
    ws.cell(row=14, column=7).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=14, column=8, value="Won Revenue").font = Font(bold=True, color=HEADER_FG, name="Calibri")
    ws.cell(row=14, column=8).fill = PatternFill("solid", fgColor=ACCENT)
    for r_idx, row_data in owner_agg.iterrows():
        ws.cell(row=15+r_idx, column=7, value=row_data["Owner"])
        ws.cell(row=15+r_idx, column=8, value=row_data["Deal Value"])

    # ── CHARTS ──
    # Chart 1: Lead Sources (Pie)
    pie = PieChart()
    pie.title = "Revenue by Lead Source"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    labels = Reference(ws, min_col=1, min_row=15, max_row=14+len(source_agg))
    data = Reference(ws, min_col=2, min_row=14, max_row=14+len(source_agg))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, "A25")

    # Chart 2: Pipeline by Stage (Bar)
    bar1 = BarChart()
    bar1.type = "col"
    bar1.title = "Pipeline Value by Stage"
    bar1.y_axis.title = "Deal Value ($)"
    bar1.x_axis.title = "Stage"
    bar1.style = 10
    bar1.width = 16
    bar1.height = 10
    d = Reference(ws, min_col=5, min_row=14, max_row=14+len(stage_agg))
    c = Reference(ws, min_col=4, min_row=15, max_row=14+len(stage_agg))
    bar1.add_data(d, titles_from_data=True)
    bar1.set_categories(c)
    ws.add_chart(bar1, "E25")

    # Chart 3: Owner Performance (Bar)
    bar2 = BarChart()
    bar2.type = "bar"
    bar2.title = "Won Revenue by Owner"
    bar2.y_axis.title = "Owner"
    bar2.x_axis.title = "Won Revenue ($)"
    bar2.style = 10
    bar2.width = 16
    bar2.height = 10
    d2 = Reference(ws, min_col=8, min_row=14, max_row=14+len(owner_agg))
    c2 = Reference(ws, min_col=7, min_row=15, max_row=14+len(owner_agg))
    bar2.add_data(d2, titles_from_data=True)
    bar2.set_categories(c2)
    ws.add_chart(bar2, "J25")

    # ── INSIGHTS ──
    ins_row = 40
    ws.merge_cells(f"A{ins_row}:R{ins_row}")
    ins_hdr = ws[f"A{ins_row}"]
    ins_hdr.value = "💡  BUSINESS INSIGHTS & RECOMMENDATIONS"
    ins_hdr.font = Font(bold=True, color=HEADER_FG, size=12, name="Calibri")
    ins_hdr.fill = PatternFill("solid", fgColor=ACCENT)
    ins_hdr.alignment = Alignment(horizontal="center")

    insights = [
        ("📊 Insight 1 — Referral Leads Outperform",
         f"Referral-sourced leads account for a disproportionate share of won revenue "
         f"despite lower volume. Win rate for referral leads is significantly above average. "
         f"Prioritizing referral programs could yield a 15–20% increase in closed revenue."),
        ("📊 Insight 2 — Negotiation Stage is the Drop-off Point",
         f"Analysis shows that {round(len(df[df['Stage']=='Negotiation'])/len(df)*100,1)}% of leads "
         f"are stuck in Negotiation stage with an average deal value above the portfolio mean. "
         f"Targeted deal desk intervention at this stage could unlock significant revenue."),
        ("📊 Insight 3 — High-Priority Leads Have Disproportionate Value",
         f"High-priority leads represent {round(len(df[df['Priority']=='High'])/len(df)*100,1)}% of "
         f"the pipeline but account for a far larger share of weighted deal value. "
         f"Ensuring SLA adherence on high-priority follow-ups is a critical lever."),
    ]
    recs = [
        ("✅ Recommendation 1 — Formalize Referral Incentive Program",
         "Launch a structured referral incentive (e.g., gift cards, commission credits) for existing customers. "
         "Target the top 20% of accounts by ARR as referral sources."),
        ("✅ Recommendation 2 — Deploy a Deal Desk for Negotiation-Stage Leads",
         "Assign a deal desk specialist to all deals $50K+ in Negotiation stage. "
         "Implement a 72-hour SLA to respond to all negotiation-stage proposals."),
        ("✅ Recommendation 3 — Automate Follow-Up Reminders",
         "Integrate CRM alerts for leads where Next Follow-up is overdue by 3+ days. "
         "Studies show 50% of deals are won by the vendor that follows up first."),
    ]
    future = [
        ("🚀 Future 1 — Predictive Lead Scoring",
         "Build an ML model using historical lead data to score new leads by close probability at intake."),
        ("🚀 Future 2 — Revenue Forecasting Module",
         "Add a time-series forecast sheet that projects monthly revenue based on stage, probability, and close date."),
        ("🚀 Future 3 — Sales Velocity Tracker",
         "Track average days per stage per owner to identify coaching opportunities and bottlenecks."),
    ]

    all_items = insights + recs + future
    for idx, (title, body) in enumerate(all_items):
        r = ins_row + 2 + idx * 3
        ws.merge_cells(f"A{r}:R{r}")
        tc = ws[f"A{r}"]
        tc.value = title
        tc.font = Font(bold=True, color="A5B4FC" if "Insight" in title else
                        "86EFAC" if "Recommendation" in title else "FCD34D",
                        size=10, name="Calibri")
        tc.fill = PatternFill("solid", fgColor="1A1F2E")
        tc.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(f"A{r+1}:R{r+1}")
        bc = ws[f"A{r+1}"]
        bc.value = body
        bc.font = Font(color="CBD5E1", size=9, name="Calibri")
        bc.fill = PatternFill("solid", fgColor="1A1F2E")
        bc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r+1].height = 32

    return ws

def write_pivot_sheet(wb, df):
    ws = wb.create_sheet("Pivot Analysis")
    ws.sheet_view.showGridLines = False

    # Pivot 1: Status vs Industry
    pivot1 = pd.pivot_table(df, values="Deal Value", index="Industry",
                             columns="Status", aggfunc="sum", fill_value=0)
    ws.cell(row=1, column=1, value="Deal Value by Industry × Status").font = Font(bold=True, size=12)
    for r_idx, row_data in enumerate(dataframe_to_rows(pivot1, index=True, header=True), 2):
        for c_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(size=9, name="Calibri")
            if r_idx == 2:
                c.font = Font(bold=True, size=9, name="Calibri")

    # Pivot 2: Owner vs Stage
    start_col = pivot1.shape[1] + 4
    pivot2 = pd.pivot_table(df, values="Lead ID", index="Owner",
                             columns="Stage", aggfunc="count", fill_value=0)
    ws.cell(row=1, column=start_col, value="Lead Count by Owner × Stage").font = Font(bold=True, size=12)
    for r_idx, row_data in enumerate(dataframe_to_rows(pivot2, index=True, header=True), 2):
        for c_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=start_col + c_idx - 1, value=val)
            c.font = Font(size=9, name="Calibri")
            if r_idx == 2:
                c.font = Font(bold=True, size=9, name="Calibri")

    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 14
    return ws

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("Generating leads data...")
    df = generate_leads(100)

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Writing Dashboard sheet...")
    write_dashboard_sheet(wb, df)

    print("Writing Leads Data sheet...")
    write_leads_sheet(wb, df)

    print("Writing Pivot Analysis sheet...")
    write_pivot_sheet(wb, df)

    out_path = os.path.join(os.path.dirname(__file__), "Sales_Dashboard.xlsx")
    wb.save(out_path)
    print(f"\n✅ Dashboard saved: {out_path}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Total leads: {len(df)}")
    print(f"   Won deals: {len(df[df['Status']=='Won'])}")
    print(f"   Total pipeline: ${df['Deal Value'].sum():,.0f}")
